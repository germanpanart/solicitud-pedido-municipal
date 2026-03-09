"""
app_web.py — Solicitud de Pedido (Versión Web)
Municipalidad de Lomas de Zamora
Migración de Tkinter → Streamlit
"""

import io
import os
import tempfile
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                TableStyle)
from zoneinfo import ZoneInfo

# =============================================================================
# CONFIGURACIÓN DE PÁGINA
# =============================================================================
st.set_page_config(
    page_title="Solicitud de Pedido – Municipalidad de Lomas de Zamora",
    page_icon="🏛️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# =============================================================================
# ESTILOS CSS PERSONALIZADOS
# =============================================================================
st.markdown("""
<style>
    /* Fondo general */
    .stApp { background-color: #f5f7fa; }

    /* Separadores horizontales (líneas finas azules) */
    hr {
        height: 1px !important;
        background-color: #1f4e79 !important;
        border: none !important;
        margin: 15px 0 !important;
    }

    /* Encabezado principal */
    .main-header {
        background: linear-gradient(135deg, #1f4e79 0%, #2e75b6 100%);
        color: white;
        padding: 20px 24px;
        border-radius: 10px;
        margin-bottom: 15px;
        box-shadow: 0 4px 16px rgba(31,78,121,0.18);
    }
    .main-header h1 { margin: 0; font-size: 1.8rem; font-weight: 800; letter-spacing: 0.5px; }
    .main-header p  { margin: 4px 0 0; font-size: 0.95rem; opacity: 0.9; }

    /* Tarjetas de sección - MÁS COMPACTAS */
    .section-card {
        background: white;
        border-radius: 8px;
        padding: 16px 20px;
        margin-bottom: 10px;
        margin-bottom: 15px; /* Ajustado para dar aire */
        border-left: 4px solid #1f4e79;
        box-shadow: 0 2px 6px rgba(0,0,0,0.06);
    }
    .section-title {
        color: #1f4e79;
        font-size: 1.05rem;
        font-weight: 700;
        margin-bottom: 10px;
        border-bottom: 1px solid #eef2f7;
        padding-bottom: 6px;
    }

    /* Tabla de artículos */
    .articulos-table {
        width: 100%;
        border-collapse: collapse;
        font-size: 0.9rem;
        margin-top: 5px;
    }
    .articulos-table th {
        background: #1f4e79;
        color: white;
        padding: 6px 10px;
        text-align: left;
        font-weight: 600;
    }
    .articulos-table td {
        padding: 5px 10px;
        border-bottom: 1px solid #e8eef4;
    }
    .articulos-table tr:nth-child(even) td { background: #f4f8fc; }

    /* Botón principal */
    div[data-testid="stDownloadButton"] > button,
    .stButton > button.generar-btn {
        background: linear-gradient(135deg, #1f4e79, #2e75b6);
        color: white;
        font-weight: 700;
        border-radius: 6px;
        padding: 8px 30px;
        border: none;
        box-shadow: 0 3px 6px rgba(31,78,121,0.2);
    }

    /* Sidebar */
    section[data-testid="stSidebar"] { background: #1f4e79 !important; }
    section[data-testid="stSidebar"] * { color: white !important; }

    /* Inputs */
    .stSelectbox label, .stTextArea label, .stTextInput label,
    .stNumberInput label, .stRadio label { font-weight: 600 !important; color: #1f4e79 !important; margin-bottom: -5px !important; }

    /* Alerta de error */
    .stAlert { border-radius: 8px; }
</style>
""", unsafe_allow_html=True)


# =============================================================================
# DATOS ESTÁTICOS
# =============================================================================
UNIDADES_EJECUTORAS = [
    "Jefatura de Gabinete", "Secretaría de Ambiente", "Secretaría de Deportes",
    "Secretaría de Desarrollo Social", "Secretaría de Educación", "Secretaría de Gobierno",
    "Secretaría de Hábitat e Integración Comunitaria", "Secretaría de Hacienda",
    "Secretaría de Mujeres, Géneros y Diversidad", "Secretaría de Obras Públicas",
    "Secretaría de Relaciones con la Comunidad", "Secretaría de Salud",
    "Secretaría de Seguridad y Justicia", "Secretaría de Servicios Públicos",
    "Secretaría General", "Secretaría Privada", "Honorable Concejo Deliberante",
]

JURISDICCIONES = [
    "1.1.1.01.38.000 Jefatura de Gabinete",
    "1.1.1.01.31.000 Secretaría de Ambiente",
    "1.1.1.01.41.000 Secretaría de Deportes",
    "1.1.1.01.30.000 Secretaría de Desarrollo Social",
    "1.1.1.01.48.000 Secretaría de Educación",
    "1.1.1.01.28.000 Secretaría de Gobierno",
    "1.1.1.01.40.000 Secretaría de Hábitat e Integración Comunitaria",
    "1.1.1.01.02.000 Secretaría de Hacienda",
    "1.1.1.01.39.000 Secretaría de Mujeres, Géneros y Diversidad",
    "1.1.1.01.35.000 Secretaría de Obras Públicas",
    "1.1.1.01.47.000 Secretaría de Relaciones con la Comunidad",
    "1.1.1.01.29.000 Secretaría de Salud",
    "1.1.1.01.22.000 Secretaría de Seguridad y Justicia",
    "1.1.1.01.46.000 Secretaría de Servicios Públicos",
    "1.1.1.01.45.000 Secretaría General",
    "1.1.1.01.33.000 Secretaría Privada",
    "1.1.1.02.00.000 Honorable Concejo Deliberante",
]

FUENTES_FINANCIAMIENTO = [
    "1.1.0 Tesoro Municipal",
    "1.3.1 De Origen Municipal",
    "1.3.2 De Origen Provincial",
    "1.3.3 De Origen Nacional",
]

PLAZOS_ENTREGA = [
    "Dentro de los 7 días corridos (notificación O.C.)",
    "Dentro de los 15 días corridos (notificación O.C.)",
    "Dentro de los 30 días corridos (notificación O.C.)",
]


# =============================================================================
# CARGA DE CATÁLOGOS (cacheada para performance)
# =============================================================================
@st.cache_data(show_spinner="Cargando catálogos…")
def cargar_catalogos():
    """Lee los Excel de la carpeta ./catalogos/ y devuelve las listas."""
    ruta_base = Path("./catalogos")
    ruta_main = ruta_base / "BASE DE DATOS MAIN.xlsx"
    ruta_articulos = ruta_base / "ARTICULOS.xlsx"

    lista_categorias: list[str] = []
    lista_objetos: list[str] = []
    lista_articulos: list[str] = []

    if ruta_main.exists():
        try:
            with pd.ExcelFile(ruta_main, engine="openpyxl") as xls:
                df_cat = pd.read_excel(xls, sheet_name=0)
                lista_categorias = df_cat.iloc[:, 0].dropna().astype(str).unique().tolist()
                df_obj = pd.read_excel(xls, sheet_name=1)
                lista_objetos = df_obj.iloc[:, 0].dropna().astype(str).unique().tolist()
        except Exception as e:
            st.sidebar.warning(f"⚠️ No se pudo leer BASE DE DATOS MAIN: {e}")

    if ruta_articulos.exists():
        try:
            df_art = pd.read_excel(ruta_articulos, engine="openpyxl")
            df_art.columns = df_art.columns.str.strip()
            df_art["COMBINADO"] = (
                df_art["ARTICULOS"].astype(str)
                + " - "
                + df_art["UNIDAD DE MEDIDA"].astype(str)
                + " - (Cod: "
                + df_art["CÓDIGO FUENTE"].astype(str)
                + ")"
            )
            lista_articulos = df_art["COMBINADO"].dropna().tolist()
        except Exception as e:
            st.sidebar.warning(f"⚠️ No se pudo leer ARTICULOS: {e}")
    else:
        st.sidebar.info("ℹ️ Archivo ARTICULOS.xlsx no encontrado en ./catalogos/")

    return lista_categorias, lista_objetos, lista_articulos


# =============================================================================
# GENERADOR DE PDF EN MEMORIA
# =============================================================================
def generar_pdf_bytes(datos: dict) -> bytes:
    """Genera el PDF y lo devuelve como bytes (sin escribir al disco)."""
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()

    estilo_titulo = ParagraphStyle("Titulo", parent=styles["Title"], alignment=1)
    estilo_seccion = ParagraphStyle("Seccion", parent=styles["Heading2"], spaceAfter=10)
    estilo_campo = ParagraphStyle("Campo", parent=styles["Normal"], fontName="Helvetica-Bold")
    estilo_tabla = ParagraphStyle(
        "TablaArticulos",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=11,
        wordWrap="CJK",
    )

    contenido = []

    # Fecha de generación
    fecha_generacion = datos.get("Fecha de generación", "")
    contenido.append(
        Paragraph(
            f"<para alignment='right'><font size=9>Fecha de generación: {fecha_generacion}</font></para>",
            styles["Normal"]
        )
    )
    contenido.append(Spacer(1, 10))

    # TÍTULO
    contenido.append(Paragraph("FORMULARIO DE SOLICITUD DE PEDIDO (SP)", estilo_titulo))
    contenido.append(Spacer(1, 20))

    # Helper para tablas de campos
    def crear_tabla_campos(campos):
        filas = []
        for campo in campos:
            valor = datos.get(campo, "-")
            if isinstance(valor, list):
                valor = ", ".join(str(v) for v in valor)
            filas.append([
                Paragraph(campo, estilo_campo),
                Paragraph(str(valor) if valor else "-", styles["Normal"]),
            ])
        tabla = Table(filas, colWidths=[7 * cm, 9 * cm])
        tabla.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.7, colors.black),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        return tabla

    # SECCIÓN 1: Datos Presupuestarios
    contenido.append(Paragraph("1. Datos Presupuestarios", estilo_seccion))
    contenido.append(crear_tabla_campos([
        "Unidad Ejecutora",
        "Jurisdicción",
        "Fuente de Financiamiento",
        "Categoría Programática",
        "Ejercicio / Plurianual",
    ]))
    contenido.append(Spacer(1, 20))

    # SECCIÓN 2: Datos de la Contratación
    contenido.append(Paragraph("2. Datos de la Contratación", estilo_seccion))
    contenido.append(crear_tabla_campos([
        "Área / Oficina solicitante",
        "Objeto de la contratación o adquisición",
        "Objeto de la contratación o adquisición_aclaracion",
        "Período de consumo / contratación",
        "Lugar y dirección de entrega",
        "Plazo de entrega",
        "Antecedentes de carga",
        "Observaciones",
    ]))
    contenido.append(Spacer(1, 20))

    # SECCIÓN 3: Artículos
    contenido.append(Paragraph("3. Artículos solicitados", estilo_seccion))
    articulos = datos.get("Seleccionar Artículo (Detalle - U.M. - Código)", [])

    tabla_art = [[
        Paragraph("N°", estilo_tabla),
        Paragraph("Detalle", estilo_tabla),
        Paragraph("U.M.", estilo_tabla),
        Paragraph("Código", estilo_tabla),
        Paragraph("Cantidad", estilo_tabla),
    ]]

    if isinstance(articulos, list) and articulos:
        for i, art in enumerate(articulos, 1):
            detalle = art.get("detalle", "-")
            cantidad = art.get("cantidad", "-")
            unidad = "-"
            codigo = "-"
            if " - " in detalle:
                partes = detalle.split(" - ")
                if len(partes) >= 3:
                    detalle_str = partes[0]
                    unidad = partes[1]
                    codigo = partes[2].replace("(Cod:", "").replace(")", "").strip()
                else:
                    detalle_str = detalle
            else:
                detalle_str = detalle

            tabla_art.append([
                Paragraph(str(i), estilo_tabla),
                Paragraph(detalle_str, estilo_tabla),
                Paragraph(unidad, estilo_tabla),
                Paragraph(codigo, estilo_tabla),
                Paragraph(str(cantidad), estilo_tabla),
            ])
    else:
        tabla_art.append([
            Paragraph("-", estilo_tabla),
            Paragraph("No se cargaron artículos", estilo_tabla),
            Paragraph("-", estilo_tabla),
            Paragraph("-", estilo_tabla),
            Paragraph("-", estilo_tabla),
        ])

    t = Table(tabla_art, colWidths=[1.5 * cm, 9 * cm, 2 * cm, 3 * cm, 2 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("BOX", (0, 0), (-1, -1), 0.7, colors.black),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (2, 1), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    contenido.append(t)
    contenido.append(Spacer(1, 30))

    # FIRMA
    firma = Table(
        [["____________________________"], ["Firma y aclaración del solicitante"]],
        colWidths=[8 * cm],
    )
    firma.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    contenido.append(firma)

    doc.build(contenido)
    buffer.seek(0)
    return buffer.getvalue()


# =============================================================================
# GENERADOR DE EXCEL EN MEMORIA
# =============================================================================

MAPEO_ENCABEZADOS = {
    "Objeto de la contratación o adquisición (Especificar qué se necesita de manera general, para qué y el período aproximado de consumo o prestación)":
        "Objeto de la contratación o adquisición",
    "Seleccionar Artículo (Detalle - U.M. - Código)":
        "Artículos Solicitados",
}


def generar_excel_bytes(datos: dict) -> bytes:
    """
    Genera el Excel acumulativo y lo devuelve como bytes.
    Si existe base_solicitudes.xlsx en disco, agrega la fila allí también
    (comportamiento acumulativo original) y devuelve los bytes del archivo actualizado.
    """
    ARCHIVO_EXCEL = "base_solicitudes.xlsx"

    if os.path.exists(ARCHIVO_EXCEL):
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Solicitudes"

        # Encabezados (solo primera vez)
        claves = [k for k in datos.keys() if not k.endswith("_EXCEL")]
        if "Fecha de generación" not in claves:
            claves.append("Fecha de generación")

        encabezados = [MAPEO_ENCABEZADOS.get(campo, campo) for campo in claves]
        ws.append(encabezados)

    encabezados_excel = [cell.value for cell in ws[1]]

    fila = []
    for encabezado_excel in encabezados_excel:
        campo_original = None
        for k, v in MAPEO_ENCABEZADOS.items():
            if v == encabezado_excel:
                campo_original = k
                break

        clave_a_buscar = campo_original or encabezado_excel

        if clave_a_buscar == "Seleccionar Artículo (Detalle - U.M. - Código)":
            valor = datos.get(f"{clave_a_buscar}_EXCEL", "")
            if not valor and isinstance(datos.get(clave_a_buscar), list):
                valor = " - ".join(
                    i.get("detalle", "-") for i in datos.get(clave_a_buscar, [])
                )
        else:
            valor = datos.get(clave_a_buscar, "")

        if encabezado_excel == "Fecha de generación":
            fila.append(datos.get("Fecha de generación", ""))
        else:
            fila.append(valor)

    ws.append(fila)

    # Guardar a disco (acumulativo) y a memoria (descarga)
    wb.save(ARCHIVO_EXCEL)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# =============================================================================
# INICIALIZACIÓN DEL SESSION STATE
# =============================================================================
if "articulos_lista" not in st.session_state:
    st.session_state.articulos_lista = []

if "mostrar_descargas" not in st.session_state:
    st.session_state.mostrar_descargas = False

if "pdf_bytes" not in st.session_state:
    st.session_state.pdf_bytes = None

if "excel_bytes" not in st.session_state:
    st.session_state.excel_bytes = None

if "nombre_archivo" not in st.session_state:
    st.session_state.nombre_archivo = ""


# =============================================================================
# SIDEBAR
# =============================================================================
with st.sidebar:
    # Logo
    logo_path = Path("./logo munipa.jpg")
    if logo_path.exists():
        st.image(str(logo_path), use_container_width=True)
    else:
        st.markdown("### 🏛️ Municipalidad\nde Lomas de Zamora")

    st.markdown("---")
    st.markdown("**Sistema de Solicitud de Pedido**")
    st.markdown("Formulario de carga inicial para trámites licitatorios.")
    st.markdown("---")
    st.markdown("**Instrucciones:**")
    st.markdown("""
1. Complete los datos presupuestarios.
2. Complete los datos de la contratación.
3. Busque y añada los artículos.
4. Presione **GENERAR** para crear PDF y Excel.
5. Descargue los archivos generados.
""")
    st.markdown("---")
    st.caption("v2.0 · Prototipo Web")


# =============================================================================
# ENCABEZADO PRINCIPAL
# =============================================================================
st.markdown("""
<div class="main-header">
    <h1>📋 SOLICITUD DE PEDIDO</h1>
    <p>Formulario de carga inicial para trámites licitatorios — Municipalidad de Lomas de Zamora</p>
</div>
""", unsafe_allow_html=True)


# =============================================================================
# CARGA DE CATÁLOGOS
# =============================================================================
categorias_excel, objetos_excel, articulos_excel = cargar_catalogos()


# =============================================================================
# SECCIÓN 1 — DATOS PRESUPUESTARIOS
# =============================================================================
st.markdown('<div class="section-card">', unsafe_allow_html=True)
st.markdown('<div class="section-title">📂 1. Datos Presupuestarios</div>', unsafe_allow_html=True)

col1, col2 = st.columns(2)

with col1:
    unidad_ejecutora = st.selectbox(
        "Unidad Ejecutora *",
        options=["— Seleccionar —"] + UNIDADES_EJECUTORAS,
        key="unidad_ejecutora",
    )
    fuente_financiamiento = st.selectbox(
        "Fuente de Financiamiento",
        options=["— Seleccionar —"] + FUENTES_FINANCIAMIENTO,
        key="fuente_financiamiento",
    )

with col2:
    jurisdiccion = st.selectbox(
        "Jurisdicción",
        options=["— Seleccionar —"] + JURISDICCIONES,
        key="jurisdiccion",
    )

    categoria_programatica_select = st.selectbox(
    "Categoría Programática (seleccionar de lista)",
    options=["— Seleccionar —"] + categorias_excel if categorias_excel else ["— Seleccionar —"],
    key="categoria_programatica",
)

categoria_programatica_manual = st.text_input(
    "Categoría Programática (escribir manualmente si no está en la lista)",
    placeholder="Ej: 01.02.03 Programa de mantenimiento urbano…",
    key="categoria_programatica_manual",
)

ejercicio = st.radio(
    "Ejercicio / Plurianual *",
    options=["2026", "Plurianual 2027"],
    horizontal=True,
    key="ejercicio",
)

st.markdown("</div>", unsafe_allow_html=True)


# =============================================================================
# SECCIÓN 2 — DATOS DE LA CONTRATACIÓN
# =============================================================================
st.markdown('<div class="section-card">', unsafe_allow_html=True)
st.markdown('<div class="section-title">📝 2. Datos de la Contratación</div>', unsafe_allow_html=True)

col3, col4 = st.columns(2)

with col3:
    area_oficina = st.text_area(
        "Área / Oficina solicitante",
        height=80,
        placeholder="Ej: Dirección de Compras, piso 3…",
        key="area_oficina",
    )

    if objetos_excel:
        objeto_contratacion = st.selectbox(
            "Objeto de la contratación o adquisición",
            options=["— Seleccionar —"] + objetos_excel,
            key="objeto_contratacion",
        )
    else:
        objeto_contratacion = st.text_input(
            "Objeto de la contratación o adquisición",
            placeholder="Ingrese el objeto manualmente…",
            key="objeto_contratacion_manual",
        )

    objeto_aclaracion = st.text_area(
        "Aclaración / Detalle de lo que se necesita",
        height=80,
        placeholder="Especifique qué se necesita, para qué y el período aproximado…",
        key="objeto_aclaracion",
    )

with col4:
    periodo_consumo = st.text_area(
        "Período de consumo / contratación",
        height=80,
        placeholder="Ej: Enero – Diciembre 2026…",
        key="periodo_consumo",
    )
    lugar_entrega = st.text_area(
        "Lugar y dirección de entrega",
        height=80,
        placeholder="Ej: Av. Hipólito Yrigoyen 3863, Lomas de Zamora…",
        key="lugar_entrega",
    )

plazo_entrega = st.radio(
    "Plazo de entrega",
    options=PLAZOS_ENTREGA,
    key="plazo_entrega",
)

col5, col6 = st.columns(2)
with col5:
    antecedentes = st.text_area(
        "Antecedentes de carga",
        height=80,
        placeholder="Número de expediente, contrataciones anteriores…",
        key="antecedentes",
    )
with col6:
    observaciones = st.text_area(
        "Observaciones",
        height=80,
        placeholder="Cualquier información adicional relevante…",
        key="observaciones",
    )

st.markdown("</div>", unsafe_allow_html=True)


# =============================================================================
# SECCIÓN 3 — ARTÍCULOS SOLICITADOS
# =============================================================================
st.markdown('<div class="section-card">', unsafe_allow_html=True)
st.markdown('<div class="section-title">📦 3. Artículos Solicitados</div>', unsafe_allow_html=True)

# Sub-fila de búsqueda + cantidad + botón
col_busq, col_cant, col_btn = st.columns([5, 1.5, 1.5])

with col_busq:
    if articulos_excel:
        articulo_seleccionado = st.selectbox(
            "Buscar artículo (escriba para filtrar)",
            options=["— Buscar artículo —"] + articulos_excel,
            key="articulo_buscador",
        )
    else:
        articulo_seleccionado = st.text_input(
            "Artículo (ingrese manualmente)",
            placeholder="Detalle - U.M. - (Cod: …)",
            key="articulo_manual",
        )

with col_cant:
    cantidad_art = st.number_input(
        "Cantidad",
        min_value=1,
        max_value=1_000_000,
        value=1,
        step=1,
        key="cantidad_articulo",
    )

with col_btn:
    st.markdown("<br>", unsafe_allow_html=True)  # Alinear verticalmente
    if st.button("➕ AÑADIR", use_container_width=True, key="btn_añadir"):
        articulo_val = articulo_seleccionado if articulos_excel else st.session_state.get("articulo_manual", "")
        if articulo_val and articulo_val != "— Buscar artículo —":
            st.session_state.articulos_lista.append({
                "detalle": articulo_val,
                "cantidad": int(cantidad_art),
            })
            st.success(f"✅ Artículo añadido: {articulo_val[:60]}…" if len(articulo_val) > 60 else f"✅ Artículo añadido.")
        else:
            st.warning("⚠️ Seleccione o ingrese un artículo antes de añadir.")

# Tabla dinámica de artículos añadidos
if st.session_state.articulos_lista:
    st.markdown("**Artículos añadidos a la solicitud:**")

    filas_html = ""
    for i, art in enumerate(st.session_state.articulos_lista):
        detalle = art["detalle"]
        unidad, codigo = "-", "-"
        if " - " in detalle:
            partes = detalle.split(" - ")
            if len(partes) >= 3:
                detalle_corto = partes[0]
                unidad = partes[1]
                codigo = partes[2].replace("(Cod:", "").replace(")", "").strip()
            else:
                detalle_corto = detalle
        else:
            detalle_corto = detalle

        filas_html += f"""
        <tr>
            <td>{i + 1}</td>
            <td>{detalle_corto}</td>
            <td style="text-align:center">{unidad}</td>
            <td style="text-align:center">{codigo}</td>
            <td style="text-align:center">{art['cantidad']}</td>
        </tr>"""

    st.markdown(f"""
    <table class="articulos-table">
        <thead>
            <tr>
                <th>N°</th>
                <th>Detalle</th>
                <th>U.M.</th>
                <th>Código</th>
                <th>Cantidad</th>
            </tr>
        </thead>
        <tbody>{filas_html}</tbody>
    </table>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Botón para quitar el último artículo (o uno específico por número)
    col_quitar_idx, col_quitar_btn = st.columns([2, 2])
    with col_quitar_idx:
        idx_quitar = st.number_input(
            "N° de artículo a quitar",
            min_value=1,
            max_value=len(st.session_state.articulos_lista),
            value=len(st.session_state.articulos_lista),
            step=1,
            key="idx_quitar",
        )
    with col_quitar_btn:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🗑️ Quitar artículo", use_container_width=True, key="btn_quitar"):
            st.session_state.articulos_lista.pop(idx_quitar - 1)
            st.rerun()

    if st.button("🗑️ Limpiar todos los artículos", key="btn_limpiar_todos"):
        st.session_state.articulos_lista = []
        st.rerun()
else:
    st.info("ℹ️ Aún no se añadieron artículos. Use el buscador de arriba.")

st.markdown("</div>", unsafe_allow_html=True)


# =============================================================================
# BOTÓN GENERAR
# =============================================================================
st.markdown("---")
col_gen_l, col_gen_c, col_gen_r = st.columns([2, 3, 2])

with col_gen_c:
    generar_clicked = st.button(
        "🖨️  GENERAR SOLICITUD",
        use_container_width=True,
        key="btn_generar",
        type="primary",
    )

if generar_clicked:
    # ── Validaciones ──────────────────────────────────────────────────────────
    errores = []

    unidad_val = st.session_state.get("unidad_ejecutora", "— Seleccionar —")
    if not unidad_val or unidad_val == "— Seleccionar —":
        errores.append("La **Unidad Ejecutora** es obligatoria.")

    if not st.session_state.articulos_lista:
        errores.append("Debe añadir **al menos un artículo** a la solicitud.")

    ejercicio_val = st.session_state.get("ejercicio", "")
    if not ejercicio_val:
        errores.append("Debe seleccionar el **Ejercicio / Plurianual**.")

    if errores:
        for err in errores:
            st.error(f"❌ {err}")
    else:
        # ── Construcción del diccionario de datos ──────────────────────────────
        nombre_campo_art = "Seleccionar Artículo (Detalle - U.M. - Código)"

        cat_select = st.session_state.get("categoria_programatica", "— Seleccionar —")
        cat_manual = st.session_state.get("categoria_programatica_manual", "").strip()
        if cat_manual:
            cat_val = cat_manual
        elif cat_select != "— Seleccionar —":
            cat_val = cat_select
        else:
            cat_val = ""

        obj_val = (
            st.session_state.get("objeto_contratacion", "— Seleccionar —")
            if objetos_excel
            else st.session_state.get("objeto_contratacion_manual", "")
        )
        if obj_val == "— Seleccionar —":
            obj_val = ""

        juris_val = st.session_state.get("jurisdiccion", "— Seleccionar —")
        if juris_val == "— Seleccionar —":
            juris_val = ""

        fuente_val = st.session_state.get("fuente_financiamiento", "— Seleccionar —")
        if fuente_val == "— Seleccionar —":
            fuente_val = ""

        now = datetime.now(ZoneInfo("America/Argentina/Buenos_Aires"))

        fecha_generacion = now.strftime("%d/%m/%Y %H:%M")
        timestamp = now.strftime("%Y%m%d_%H%M%S")

        datos = {
            "Fecha de generación": fecha_generacion,
            "Unidad Ejecutora": unidad_val,
            "Jurisdicción": juris_val,
            "Fuente de Financiamiento": fuente_val,
            "Categoría Programática": cat_val,
            "Ejercicio / Plurianual": ejercicio_val,
            "Área / Oficina solicitante": st.session_state.get("area_oficina", ""),
            nombre_campo_art: st.session_state.articulos_lista,
            f"{nombre_campo_art}_EXCEL": " - ".join(
                i.get("detalle", "-") for i in st.session_state.articulos_lista
            ),
            "Objeto de la contratación o adquisición": obj_val,
            "Objeto de la contratación o adquisición_aclaracion": st.session_state.get("objeto_aclaracion", ""),
            "Período de consumo / contratación": st.session_state.get("periodo_consumo", ""),
            "Lugar y dirección de entrega": st.session_state.get("lugar_entrega", ""),
            "Plazo de entrega": st.session_state.get("plazo_entrega", ""),
            "Antecedentes de carga": st.session_state.get("antecedentes", ""),
            "Observaciones": st.session_state.get("observaciones", ""),
        }

        # ── Generación ────────────────────────────────────────────────────────
        nombre_base = f"Solicitud_Pedido_{timestamp}"

        with st.spinner("Generando documentos…"):
            try:
                pdf_bytes = generar_pdf_bytes(datos)
                excel_bytes = generar_excel_bytes(datos)

                st.session_state.pdf_bytes = pdf_bytes
                st.session_state.excel_bytes = excel_bytes
                st.session_state.nombre_archivo = nombre_base
                st.session_state.mostrar_descargas = True

            except Exception as e:
                st.error(f"❌ Error inesperado al generar los documentos: {e}")
                st.session_state.mostrar_descargas = False


# =============================================================================
# ÁREA DE DESCARGAS
# =============================================================================
if st.session_state.mostrar_descargas and st.session_state.pdf_bytes:
    st.markdown("---")
    st.success("✅ ¡Documentos generados correctamente! Descárguelos a continuación:")

    col_dl1, col_dl2, col_dl_spacer = st.columns([2, 2, 3])

    with col_dl1:
        st.download_button(
            label="📄 Descargar PDF",
            data=st.session_state.pdf_bytes,
            file_name=f"{st.session_state.nombre_archivo}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )

    with col_dl2:
        st.download_button(
            label="📊 Descargar Excel",
            data=st.session_state.excel_bytes,
            file_name=f"{st.session_state.nombre_archivo}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.markdown("---")
    st.markdown(
        "<small>ℹ️ El Excel también fue guardado en <code>base_solicitudes.xlsx</code> "
        "en la carpeta del proyecto para su acumulación.</small>",
        unsafe_allow_html=True,
    )
